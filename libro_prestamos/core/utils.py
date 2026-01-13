"""
Utilidades para el procesamiento de archivos Excel y carga masiva de libros.
"""
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from django.db.models import Q
from django.core.mail import send_mail, EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from .models import Libro
import traceback
import logging

logger = logging.getLogger(__name__)


def normalizar_nombre_columna(nombre):
    """
    Normaliza el nombre de una columna para hacer comparaciones flexibles.
    Elimina espacios, convierte a minúsculas y normaliza acentos básicos.
    """
    if not nombre:
        return ''
    # Convertir a minúsculas y eliminar espacios
    nombre = str(nombre).lower().strip()
    # Normalizar acentos básicos
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ñ': 'n'
    }
    for old, new in reemplazos.items():
        nombre = nombre.replace(old, new)
    return nombre


def encontrar_indices_columnas(worksheet):
    """
    Encuentra los índices de las columnas requeridas en la primera fila.
    Retorna un diccionario con los índices encontrados.
    """
    indices = {
        'nombre': None,
        'autor': None,
        'editorial': None,
        'isbn': None,
        'descripcion': None
    }
    
    # Leer primera fila
    primera_fila = [cell.value for cell in worksheet[1]]
    
    # Buscar columnas (case-insensitive y flexible)
    for idx, valor in enumerate(primera_fila, start=1):
        if valor is None:
            continue
        valor_normalizado = normalizar_nombre_columna(valor)
        
        # Buscar coincidencias
        if valor_normalizado in ['nombre', 'titulo', 'title', 'libro']:
            indices['nombre'] = idx
        elif valor_normalizado in ['autor', 'author', 'escritor']:
            indices['autor'] = idx
        elif valor_normalizado in ['editorial', 'publisher', 'editora']:
            indices['editorial'] = idx
        elif valor_normalizado == 'isbn':
            indices['isbn'] = idx
        elif valor_normalizado in ['descripcion', 'description', 'comentario', 'comentarios']:
            indices['descripcion'] = idx
    
    return indices


def validar_fila_libro(fila, indices, numero_fila):
    """
    Valida una fila de datos y retorna un diccionario con los datos validados
    o un error si la validación falla.
    """
    errores = []
    datos = {}
    
    # Validar campos obligatorios
    nombre = None
    autor = None
    
    if indices['nombre']:
        nombre = fila[indices['nombre'] - 1].value if fila[indices['nombre'] - 1].value else None
    if indices['autor']:
        autor = fila[indices['autor'] - 1].value if fila[indices['autor'] - 1].value else None
    
    # Validar que nombre y autor existan
    if not nombre or (isinstance(nombre, str) and not nombre.strip()):
        errores.append(f"Fila {numero_fila}: El campo 'Nombre' es obligatorio")
    else:
        nombre = str(nombre).strip()
        if len(nombre) > 255:
            errores.append(f"Fila {numero_fila}: El nombre excede los 255 caracteres")
        else:
            datos['nombre'] = nombre
    
    if not autor or (isinstance(autor, str) and not autor.strip()):
        errores.append(f"Fila {numero_fila}: El campo 'Autor' es obligatorio")
    else:
        autor = str(autor).strip()
        if len(autor) > 255:
            errores.append(f"Fila {numero_fila}: El autor excede los 255 caracteres")
        else:
            datos['autor'] = autor
    
    # Campos opcionales
    if indices['editorial']:
        editorial = fila[indices['editorial'] - 1].value
        if editorial:
            editorial = str(editorial).strip()
            if len(editorial) > 255:
                errores.append(f"Fila {numero_fila}: La editorial excede los 255 caracteres")
            else:
                datos['editorial'] = editorial
    
    if indices['isbn']:
        isbn = fila[indices['isbn'] - 1].value
        if isbn:
            isbn = str(isbn).strip()
            # Limpiar guiones y espacios del ISBN
            isbn = isbn.replace('-', '').replace(' ', '')
            if len(isbn) > 13:
                errores.append(f"Fila {numero_fila}: El ISBN excede los 13 caracteres")
            elif len(isbn) > 0:
                datos['isbn'] = isbn
    
    if indices['descripcion']:
        descripcion = fila[indices['descripcion'] - 1].value
        if descripcion:
            datos['descripcion'] = str(descripcion).strip()
    
    if errores:
        return {'error': '; '.join(errores)}
    
    return {'datos': datos}


def es_duplicado(datos_libro, usuario):
    """
    Verifica si un libro es duplicado comparando con los libros existentes del usuario.
    Primero verifica por ISBN, luego por nombre+autor.
    """
    # Verificar por ISBN si existe
    if datos_libro.get('isbn'):
        existe = Libro.objects.filter(
            propietario=usuario,
            isbn=datos_libro['isbn']
        ).exists()
        if existe:
            return True, 'ISBN'
    
    # Verificar por nombre y autor (case-insensitive)
    existe = Libro.objects.filter(
        propietario=usuario,
        nombre__iexact=datos_libro['nombre'],
        autor__iexact=datos_libro['autor']
    ).exists()
    
    if existe:
        return True, 'Nombre+Autor'
    
    return False, None


def procesar_excel_libros(archivo, usuario):
    """
    Procesa un archivo Excel y retorna los resultados de la carga masiva.
    
    Retorna un diccionario con:
    - libros_creados: lista de libros creados exitosamente
    - duplicados: lista de libros duplicados (no creados)
    - errores: lista de errores por fila
    - total_procesado: número total de filas procesadas
    """
    resultados = {
        'libros_creados': [],
        'duplicados': [],
        'errores': [],
        'total_procesado': 0
    }
    
    try:
        # Leer el archivo Excel
        workbook = load_workbook(archivo, read_only=True, data_only=True)
        worksheet = workbook.active
        
        # Encontrar índices de columnas
        indices = encontrar_indices_columnas(worksheet)
        
        # Validar que se encontraron las columnas obligatorias
        if indices['nombre'] is None or indices['autor'] is None:
            resultados['errores'].append({
                'fila': 0,
                'mensaje': 'No se encontraron las columnas obligatorias "Nombre" y "Autor" en la primera fila'
            })
            return resultados
        
        # Procesar filas (empezando desde la fila 2, ya que la 1 es encabezado)
        libros_a_crear = []
        
        for row_num, fila in enumerate(worksheet.iter_rows(min_row=2), start=2):
            resultados['total_procesado'] += 1
            
            # Validar fila
            validacion = validar_fila_libro(fila, indices, row_num)
            
            if 'error' in validacion:
                resultados['errores'].append({
                    'fila': row_num,
                    'mensaje': validacion['error']
                })
                continue
            
            datos = validacion['datos']
            
            # Verificar duplicados
            es_dup, tipo_dup = es_duplicado(datos, usuario)
            if es_dup:
                resultados['duplicados'].append({
                    'fila': row_num,
                    'libro': f"{datos['nombre']} - {datos['autor']}",
                    'tipo': tipo_dup
                })
                continue
            
            # Preparar libro para crear
            libro_data = {
                'nombre': datos['nombre'],
                'autor': datos['autor'],
                'propietario': usuario,
                'estado': 'disponible'
            }
            
            if 'editorial' in datos:
                libro_data['editorial'] = datos['editorial']
            if 'isbn' in datos:
                libro_data['isbn'] = datos['isbn']
            if 'descripcion' in datos:
                libro_data['descripcion'] = datos['descripcion']
            
            libros_a_crear.append(Libro(**libro_data))
        
        # Crear libros en lote usando bulk_create
        if libros_a_crear:
            libros_creados = Libro.objects.bulk_create(libros_a_crear)
            resultados['libros_creados'] = [
                {'nombre': libro.nombre, 'autor': libro.autor}
                for libro in libros_creados
            ]
        
        workbook.close()
        
    except Exception as e:
        resultados['errores'].append({
            'fila': 0,
            'mensaje': f'Error al procesar el archivo: {str(e)}'
        })
    
    return resultados


def generar_plantilla_excel():
    """
    Genera un archivo Excel de ejemplo con la estructura esperada.
    Retorna un objeto BytesIO con el contenido del archivo.
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Libros"
    
    # Encabezados
    ws['A1'] = 'Nombre'
    ws['B1'] = 'Autor'
    ws['C1'] = 'Editorial'
    ws['D1'] = 'ISBN'
    ws['E1'] = 'Descripción'
    
    # Estilo de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Ejemplos de datos
    ejemplos = [
        ['El Quijote', 'Miguel de Cervantes', 'Editorial Real', '9788491049000', 'Novela clásica española'],
        ['Cien años de soledad', 'Gabriel García Márquez', 'Sudamericana', '9788437604947', 'Realismo mágico'],
        ['1984', 'George Orwell', 'Seix Barral', '9788499890944', 'Distopía'],
    ]
    
    for row_num, ejemplo in enumerate(ejemplos, start=2):
        for col_num, valor in enumerate(ejemplo, start=1):
            ws.cell(row=row_num, column=col_num, value=valor)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


# Funciones de Email

def enviar_email_cambio_password(user, token):
    """
    Envía un email con el enlace para cambiar la contraseña.
    
    Args:
        user: Usuario que solicita el cambio
        token: Token de PasswordResetToken
    """
    try:
        logger.debug(f"Iniciando envío de email para usuario: {user.email}")
        # Construir URL de confirmación
        from django.urls import reverse
        
        # Obtener el dominio actual
        if settings.DEBUG:
            # En desarrollo, siempre usar 127.0.0.1:8000
            domain = '127.0.0.1:8000'
        elif settings.ALLOWED_HOSTS:
            domain = settings.ALLOWED_HOSTS[0]
        else:
            domain = '127.0.0.1:8000'
        
        # Construir URL completa
        protocol = 'https' if not settings.DEBUG else 'http'
        if not domain.startswith('http'):
            base_url = f"{protocol}://{domain}"
        else:
            base_url = domain
        
        reset_path = reverse('confirmar_cambio_password', args=[token.token])
        reset_url = f"{base_url}{reset_path}"
        
        # Renderizar template de email
        logger.debug("Renderizando template de email")
        # Usar subject sin caracteres especiales para evitar problemas con SMTP
        # En desarrollo con consola no importa, pero es buena práctica
        subject = 'Cambio de contrasena - Biblioteca Colectiva'
        try:
            html_message = render_to_string('emails/password_reset_email.html', {
                'user': user,
                'reset_url': reset_url,
                'token': token,
            })
            logger.debug(f"Template renderizado exitosamente. Longitud HTML: {len(html_message)}")
        except Exception as e:
            logger.error(f"Error al renderizar template: {e}")
            logger.error(traceback.format_exc())
            raise
        
        # Versión texto plano (opcional, para clientes que no soportan HTML)
        logger.debug("Creando mensaje de texto plano")
        plain_message = f"""
Hola {user.get_full_name() or user.username},

Has solicitado cambiar tu contraseña en Biblioteca Colectiva.

Para completar el cambio, haz clic en el siguiente enlace:
{reset_url}

Este enlace expirará en 24 horas.

Si no solicitaste este cambio, puedes ignorar este email. Tu contraseña no será modificada.

Saludos,
Equipo de Biblioteca Colectiva
"""
        
        # Usar EmailMessage para mejor manejo de UTF-8
        logger.debug(f"Creando EmailMessage. Subject: {subject[:50]}...")
        logger.debug(f"From: {settings.DEFAULT_FROM_EMAIL}, To: {user.email}")
        logger.debug(f"Encoding configurado: UTF-8")
        
        try:
            # Asegurar que el subject esté en UTF-8
            if isinstance(subject, str):
                subject = subject.encode('utf-8').decode('utf-8')
            
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[user.email],
            )
            email.content_subtype = "html"
            email.encoding = 'utf-8'
            
            # Agregar versión texto plano como alternativa
            email.alternatives = [(plain_message, 'text/plain')]
            
            logger.debug("Enviando email...")
            email.send(fail_silently=False)
            logger.debug("Email enviado exitosamente")
            return True
        except UnicodeEncodeError as ue:
            logger.error(f"Error de codificación Unicode: {ue}")
            logger.error(f"Subject: {repr(subject)}")
            logger.error(f"From: {repr(settings.DEFAULT_FROM_EMAIL)}")
            logger.error(f"To: {repr(user.email)}")
            logger.error(traceback.format_exc())
            raise
        except Exception as e:
            logger.error(f"Error al crear/enviar EmailMessage: {e}")
            logger.error(f"Tipo de error: {type(e).__name__}")
            logger.error(traceback.format_exc())
            raise
    except Exception as e:
        # Log detallado del error
        error_msg = f"Error al enviar email de cambio de contraseña: {e}"
        logger.error(error_msg)
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Usuario: {user.email if user else 'N/A'}")
        logger.error(traceback.format_exc())
        
        # También imprimir en consola para desarrollo
        print(f"\n{'='*60}")
        print(f"ERROR DETALLADO AL ENVIAR EMAIL")
        print(f"{'='*60}")
        print(f"Tipo: {type(e).__name__}")
        print(f"Mensaje: {e}")
        print(f"Usuario: {user.email if user else 'N/A'}")
        print(f"\nTraceback completo:")
        traceback.print_exc()
        print(f"{'='*60}\n")
        
        return False


def enviar_email_confirmacion_cambio(user):
    """
    Envía un email de confirmación después de cambiar la contraseña exitosamente.
    
    Args:
        user: Usuario que cambió la contraseña
    """
    try:
        subject = 'Contrasena cambiada exitosamente - Biblioteca Colectiva'
        html_message = render_to_string('emails/password_change_confirmation.html', {
            'user': user,
        })
        
        # Versión texto plano
        plain_message = f"""
Hola {user.get_full_name() or user.username},

Tu contraseña ha sido cambiada exitosamente.

Si no realizaste este cambio, por favor contacta al soporte inmediatamente.

Saludos,
Equipo de Biblioteca Colectiva
"""
        
        # Usar EmailMessage para mejor manejo de UTF-8
        # Asegurar que el subject esté en UTF-8
        if isinstance(subject, str):
            subject = subject.encode('utf-8').decode('utf-8')
        
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[user.email],
        )
        email.content_subtype = "html"
        email.encoding = 'utf-8'
        
        # Agregar versión texto plano como alternativa
        email.alternatives = [(plain_message, 'text/plain')]
        
        # Enviar email
        email.send(fail_silently=False)
        return True
    except Exception as e:
        print(f"Error al enviar email de confirmación: {e}")
        return False


def enviar_email_cambio_email(user, token, new_email):
    """
    Envía un email con el enlace para confirmar el cambio de email.
    
    Args:
        user: Usuario que solicita el cambio
        token: Token de EmailChangeToken
        new_email: Nuevo email a confirmar
    """
    try:
        logger.debug(f"Iniciando envío de email de cambio de email para usuario: {user.email}")
        # Construir URL de confirmación
        from django.urls import reverse
        
        # Obtener el dominio actual
        if settings.DEBUG:
            # En desarrollo, siempre usar 127.0.0.1:8000
            domain = '127.0.0.1:8000'
        elif settings.ALLOWED_HOSTS:
            domain = settings.ALLOWED_HOSTS[0]
        else:
            domain = '127.0.0.1:8000'
        
        # Construir URL completa
        protocol = 'https' if not settings.DEBUG else 'http'
        if not domain.startswith('http'):
            base_url = f"{protocol}://{domain}"
        else:
            base_url = domain
        
        confirm_path = reverse('confirmar_cambio_email', args=[token.token])
        confirm_url = f"{base_url}{confirm_path}"
        
        # Renderizar template de email
        logger.debug("Renderizando template de email de cambio de email")
        subject = 'Confirmar cambio de correo electronico - Biblioteca Colectiva'
        try:
            html_message = render_to_string('emails/email_change_request.html', {
                'user': user,
                'new_email': new_email,
                'confirm_url': confirm_url,
                'token': token,
            })
            logger.debug(f"Template renderizado exitosamente. Longitud HTML: {len(html_message)}")
        except Exception as e:
            logger.error(f"Error al renderizar template: {e}")
            logger.error(traceback.format_exc())
            raise
        
        # Versión texto plano
        logger.debug("Creando mensaje de texto plano")
        plain_message = f"""
Hola {user.get_full_name() or user.username},

Has solicitado cambiar tu correo electrónico en Biblioteca Colectiva.

Tu correo actual: {user.email}
Nuevo correo: {new_email}

Para completar el cambio, haz clic en el siguiente enlace:
{confirm_url}

Este enlace expirará en 24 horas.

Si no solicitaste este cambio, puedes ignorar este email. Tu correo electrónico no será modificado.

Saludos,
Equipo de Biblioteca Colectiva
"""
        
        # Usar EmailMessage para mejor manejo de UTF-8
        logger.debug(f"Creando EmailMessage. Subject: {subject[:50]}...")
        logger.debug(f"From: {settings.DEFAULT_FROM_EMAIL}, To: {new_email}")
        logger.debug(f"Encoding configurado: UTF-8")
        
        try:
            # Asegurar que el subject esté en UTF-8
            if isinstance(subject, str):
                subject = subject.encode('utf-8').decode('utf-8')
            
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[new_email],  # Enviar al nuevo email
            )
            email.content_subtype = "html"
            email.encoding = 'utf-8'
            
            # Agregar versión texto plano como alternativa
            email.alternatives = [(plain_message, 'text/plain')]
            
            logger.debug("Enviando email...")
            email.send(fail_silently=False)
            logger.debug("Email enviado exitosamente")
            return True
        except UnicodeEncodeError as ue:
            logger.error(f"Error de codificación Unicode: {ue}")
            logger.error(f"Subject: {repr(subject)}")
            logger.error(f"From: {repr(settings.DEFAULT_FROM_EMAIL)}")
            logger.error(f"To: {repr(new_email)}")
            logger.error(traceback.format_exc())
            raise
        except Exception as e:
            logger.error(f"Error al crear/enviar EmailMessage: {e}")
            logger.error(f"Tipo de error: {type(e).__name__}")
            logger.error(traceback.format_exc())
            raise
    except Exception as e:
        # Log detallado del error
        error_msg = f"Error al enviar email de cambio de email: {e}"
        logger.error(error_msg)
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Usuario: {user.email if user else 'N/A'}")
        logger.error(traceback.format_exc())
        
        # También imprimir en consola para desarrollo
        print(f"\n{'='*60}")
        print(f"ERROR DETALLADO AL ENVIAR EMAIL DE CAMBIO DE EMAIL")
        print(f"{'='*60}")
        print(f"Tipo: {type(e).__name__}")
        print(f"Mensaje: {e}")
        print(f"Usuario: {user.email if user else 'N/A'}")
        print(f"Nuevo email: {new_email if 'new_email' in locals() else 'N/A'}")
        print(f"\nTraceback completo:")
        traceback.print_exc()
        print(f"{'='*60}\n")
        
        return False


def enviar_email_confirmacion_cambio_email(user, old_email):
    """
    Envía un email de confirmación después de cambiar el email exitosamente.
    
    Args:
        user: Usuario que cambió el email
        old_email: Email anterior del usuario
    """
    try:
        subject = 'Correo electronico cambiado exitosamente - Biblioteca Colectiva'
        html_message = render_to_string('emails/email_change_confirmation.html', {
            'user': user,
            'old_email': old_email,
        })
        
        # Versión texto plano
        plain_message = f"""
Hola {user.get_full_name() or user.username},

Tu correo electrónico ha sido cambiado exitosamente.

Correo anterior: {old_email}
Correo nuevo: {user.email}

Si no realizaste este cambio, por favor contacta al soporte inmediatamente.

Saludos,
Equipo de Biblioteca Colectiva
"""
        
        # Usar EmailMessage para mejor manejo de UTF-8
        # Asegurar que el subject esté en UTF-8
        if isinstance(subject, str):
            subject = subject.encode('utf-8').decode('utf-8')
        
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[user.email],  # Enviar al nuevo email
        )
        email.content_subtype = "html"
        email.encoding = 'utf-8'
        
        # Agregar versión texto plano como alternativa
        email.alternatives = [(plain_message, 'text/plain')]
        
        # Enviar email
        email.send(fail_silently=False)
        return True
    except Exception as e:
        print(f"Error al enviar email de confirmación de cambio de email: {e}")
        return False
