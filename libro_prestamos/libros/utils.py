"""
Utilidades para el procesamiento de archivos Excel y carga masiva de libros.
"""
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from .models import Libro


def normalizar_nombre_columna(nombre):
    """
    Normaliza el nombre de una columna para hacer comparaciones flexibles.
    Elimina espacios, convierte a minúsculas y normaliza acentos básicos.
    """
    if not nombre:
        return ''
    # Convertir a minúsculas y eliminar espacios
    nombre = str(nombre).lower().strip()
    # Normalizar acentos básicos
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ñ': 'n'
    }
    for old, new in reemplazos.items():
        nombre = nombre.replace(old, new)
    return nombre


def encontrar_indices_columnas(worksheet):
    """
    Encuentra los índices de las columnas requeridas en la primera fila.
    Retorna un diccionario con los índices encontrados.
    """
    indices = {
        'nombre': None,
        'autor': None,
        'editorial': None,
        'isbn': None,
        'descripcion': None
    }
    
    # Leer primera fila
    primera_fila = [cell.value for cell in worksheet[1]]
    
    # Buscar columnas (case-insensitive y flexible)
    for idx, valor in enumerate(primera_fila, start=1):
        if valor is None:
            continue
        valor_normalizado = normalizar_nombre_columna(valor)
        
        # Buscar coincidencias
        if valor_normalizado in ['nombre', 'titulo', 'title', 'libro']:
            indices['nombre'] = idx
        elif valor_normalizado in ['autor', 'author', 'escritor']:
            indices['autor'] = idx
        elif valor_normalizado in ['editorial', 'publisher', 'editora']:
            indices['editorial'] = idx
        elif valor_normalizado == 'isbn':
            indices['isbn'] = idx
        elif valor_normalizado in ['descripcion', 'description', 'comentario', 'comentarios']:
            indices['descripcion'] = idx
    
    return indices


def validar_fila_libro(fila, indices, numero_fila):
    """
    Valida una fila de datos y retorna un diccionario con los datos validados
    o un error si la validación falla.
    """
    errores = []
    datos = {}
    
    # Validar campos obligatorios
    nombre = None
    autor = None
    
    if indices['nombre']:
        nombre = fila[indices['nombre'] - 1].value if fila[indices['nombre'] - 1].value else None
    if indices['autor']:
        autor = fila[indices['autor'] - 1].value if fila[indices['autor'] - 1].value else None
    
    # Validar que nombre y autor existan
    if not nombre or (isinstance(nombre, str) and not nombre.strip()):
        errores.append(f"Fila {numero_fila}: El campo 'Nombre' es obligatorio")
    else:
        nombre = str(nombre).strip()
        if len(nombre) > 255:
            errores.append(f"Fila {numero_fila}: El nombre excede los 255 caracteres")
        else:
            datos['nombre'] = nombre
    
    if not autor or (isinstance(autor, str) and not autor.strip()):
        errores.append(f"Fila {numero_fila}: El campo 'Autor' es obligatorio")
    else:
        autor = str(autor).strip()
        if len(autor) > 255:
            errores.append(f"Fila {numero_fila}: El autor excede los 255 caracteres")
        else:
            datos['autor'] = autor
    
    # Campos opcionales
    if indices['editorial']:
        editorial = fila[indices['editorial'] - 1].value
        if editorial:
            editorial = str(editorial).strip()
            if len(editorial) > 255:
                errores.append(f"Fila {numero_fila}: La editorial excede los 255 caracteres")
            else:
                datos['editorial'] = editorial
    
    if indices['isbn']:
        isbn = fila[indices['isbn'] - 1].value
        if isbn:
            isbn = str(isbn).strip()
            # Limpiar guiones y espacios del ISBN
            isbn = isbn.replace('-', '').replace(' ', '')
            if len(isbn) > 13:
                errores.append(f"Fila {numero_fila}: El ISBN excede los 13 caracteres")
            elif len(isbn) > 0:
                datos['isbn'] = isbn
    
    if indices['descripcion']:
        descripcion = fila[indices['descripcion'] - 1].value
        if descripcion:
            datos['descripcion'] = str(descripcion).strip()
    
    if errores:
        return {'error': '; '.join(errores)}
    
    return {'datos': datos}


def es_duplicado(datos_libro, usuario):
    """
    Verifica si un libro es duplicado comparando con los libros existentes del usuario.
    Primero verifica por ISBN, luego por nombre+autor.
    """
    # Verificar por ISBN si existe
    if datos_libro.get('isbn'):
        existe = Libro.objects.filter(
            propietario=usuario,
            isbn=datos_libro['isbn']
        ).exists()
        if existe:
            return True, 'ISBN'
    
    # Verificar por nombre y autor (case-insensitive)
    existe = Libro.objects.filter(
        propietario=usuario,
        nombre__iexact=datos_libro['nombre'],
        autor__iexact=datos_libro['autor']
    ).exists()
    
    if existe:
        return True, 'Nombre+Autor'
    
    return False, None


def procesar_excel_libros(archivo, usuario):
    """
    Procesa un archivo Excel y retorna los resultados de la carga masiva.
    
    Retorna un diccionario con:
    - libros_creados: lista de libros creados exitosamente
    - duplicados: lista de libros duplicados (no creados)
    - errores: lista de errores por fila
    - total_procesado: número total de filas procesadas
    """
    resultados = {
        'libros_creados': [],
        'duplicados': [],
        'errores': [],
        'total_procesado': 0
    }
    
    try:
        # Leer el archivo Excel
        workbook = load_workbook(archivo, read_only=True, data_only=True)
        worksheet = workbook.active
        
        # Encontrar índices de columnas
        indices = encontrar_indices_columnas(worksheet)
        
        # Validar que se encontraron las columnas obligatorias
        if indices['nombre'] is None or indices['autor'] is None:
            resultados['errores'].append({
                'fila': 0,
                'mensaje': 'No se encontraron las columnas obligatorias "Nombre" y "Autor" en la primera fila'
            })
            return resultados
        
        # Procesar filas (empezando desde la fila 2, ya que la 1 es encabezado)
        libros_a_crear = []
        
        for row_num, fila in enumerate(worksheet.iter_rows(min_row=2), start=2):
            resultados['total_procesado'] += 1
            
            # Validar fila
            validacion = validar_fila_libro(fila, indices, row_num)
            
            if 'error' in validacion:
                resultados['errores'].append({
                    'fila': row_num,
                    'mensaje': validacion['error']
                })
                continue
            
            datos = validacion['datos']
            
            # Verificar duplicados
            es_dup, tipo_dup = es_duplicado(datos, usuario)
            if es_dup:
                resultados['duplicados'].append({
                    'fila': row_num,
                    'libro': f"{datos['nombre']} - {datos['autor']}",
                    'tipo': tipo_dup
                })
                continue
            
            # Preparar libro para crear
            libro_data = {
                'nombre': datos['nombre'],
                'autor': datos['autor'],
                'propietario': usuario,
                'estado': 'disponible'
            }
            
            if 'editorial' in datos:
                libro_data['editorial'] = datos['editorial']
            if 'isbn' in datos:
                libro_data['isbn'] = datos['isbn']
            if 'descripcion' in datos:
                libro_data['descripcion'] = datos['descripcion']
            
            libros_a_crear.append(Libro(**libro_data))
        
        # Crear libros en lote usando bulk_create
        if libros_a_crear:
            libros_creados = Libro.objects.bulk_create(libros_a_crear)
            resultados['libros_creados'] = [
                {'nombre': libro.nombre, 'autor': libro.autor}
                for libro in libros_creados
            ]
        
        workbook.close()
        
    except Exception as e:
        resultados['errores'].append({
            'fila': 0,
            'mensaje': f'Error al procesar el archivo: {str(e)}'
        })
    
    return resultados


def generar_plantilla_excel():
    """
    Genera un archivo Excel de ejemplo con la estructura esperada.
    Retorna un objeto BytesIO con el contenido del archivo.
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Libros"
    
    # Encabezados
    ws['A1'] = 'Nombre'
    ws['B1'] = 'Autor'
    ws['C1'] = 'Editorial'
    ws['D1'] = 'ISBN'
    ws['E1'] = 'Descripción'
    
    # Estilo de encabezados
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Ejemplos de datos
    ejemplos = [
        ['El Quijote', 'Miguel de Cervantes', 'Editorial Real', '9788491049000', 'Novela clásica española'],
        ['Cien años de soledad', 'Gabriel García Márquez', 'Sudamericana', '9788437604947', 'Realismo mágico'],
        ['1984', 'George Orwell', 'Seix Barral', '9788499890944', 'Distopía'],
    ]
    
    for row_num, ejemplo in enumerate(ejemplos, start=2):
        for col_num, valor in enumerate(ejemplo, start=1):
            ws.cell(row=row_num, column=col_num, value=valor)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
