"""
Configuración compartida para todos los tests
"""
import pytest
from django.contrib.auth.models import User
from core.models import Perfil, PasswordResetToken
from django.utils import timezone
from datetime import timedelta


@pytest.fixture
def user(db):
    """Crea un usuario de prueba"""
    user = User.objects.create_user(
        username='testuser',
        email='test@example.com',
        password='testpass123',
        first_name='Test',
        last_name='User'
    )
    # Crear perfil asociado
    Perfil.objects.create(
        usuario=user,
        ciudad='Buenos Aires',
        pais='Argentina'
    )
    return user


@pytest.fixture
def another_user(db):
    """Crea otro usuario de prueba"""
    user = User.objects.create_user(
        username='otheruser',
        email='other@example.com',
        password='otherpass123'
    )
    Perfil.objects.create(
        usuario=user,
        ciudad='Córdoba',
        pais='Argentina'
    )
    return user


@pytest.fixture
def password_reset_token(db, user):
    """Crea un token de cambio de contraseña válido"""
    return PasswordResetToken.create_token(user)


@pytest.fixture
def expired_token(db, user):
    """Crea un token expirado"""
    token = PasswordResetToken.create_token(user)
    token.expires_at = timezone.now() - timedelta(hours=1)
    token.save()
    return token


@pytest.fixture
def used_token(db, user):
    """Crea un token ya usado"""
    token = PasswordResetToken.create_token(user)
    token.mark_as_used()
    return token


@pytest.fixture
def excel_file_valido():
    """Crea un archivo Excel válido con libros de prueba"""
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    
    # Encabezados
    ws.append(['Nombre', 'Autor', 'Editorial', 'ISBN', 'Descripción'])
    
    # Datos de prueba
    ws.append(['El Quijote', 'Miguel de Cervantes', 'Editorial Test', '9781234567890', 'Clásico de la literatura'])
    ws.append(['1984', 'George Orwell', 'Editorial Test 2', '9780987654321', 'Novela distópica'])
    ws.append(['Cien años de soledad', 'Gabriel García Márquez', None, None, None])
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


@pytest.fixture
def excel_file_con_duplicados():
    """Crea un archivo Excel con libros duplicados"""
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    
    # Encabezados
    ws.append(['Nombre', 'Autor', 'ISBN'])
    
    # Datos con duplicados
    ws.append(['Libro Original', 'Autor Original', '9781111111111'])
    ws.append(['Libro Duplicado ISBN', 'Otro Autor', '9781111111111'])  # Mismo ISBN
    ws.append(['Libro Original', 'Autor Original', None])  # Mismo nombre+autor
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


@pytest.fixture
def excel_file_con_errores():
    """Crea un archivo Excel con errores de validación"""
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    
    # Encabezados
    ws.append(['Nombre', 'Autor'])
    
    # Datos con errores
    ws.append(['', 'Autor Válido'])  # Nombre vacío
    ws.append(['Nombre Válido', ''])  # Autor vacío
    ws.append(['Nombre Válido', 'Autor Válido'])  # Fila válida
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


@pytest.fixture
def excel_file_sin_columnas_obligatorias():
    """Crea un archivo Excel sin columnas obligatorias"""
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    
    # Encabezados incorrectos
    ws.append(['Título', 'Escritor'])  # No encuentra "Nombre" ni "Autor"
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
