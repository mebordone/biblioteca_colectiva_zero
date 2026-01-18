"""
Tests para la funcionalidad de carga masiva de libros desde archivos Excel.
"""
import pytest
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from io import BytesIO
from openpyxl import Workbook
from libros.models import Libro
from libros.utils import procesar_excel_libros, generar_plantilla_excel
from libros.views import cargar_libros_masivo, descargar_plantilla_excel


class TestProcesarExcelLibros:
    """Tests para la función procesar_excel_libros()"""
    
    def test_procesamiento_exitoso(self, excel_file_valido, user):
        """Test procesamiento exitoso de archivo Excel válido"""
        resultados = procesar_excel_libros(excel_file_valido, user)
        
        assert resultados['total_procesado'] == 3
        assert len(resultados['libros_creados']) == 3
        assert len(resultados['duplicados']) == 0
        assert len(resultados['errores']) == 0
        
        # Verificar que los libros se crearon en la base de datos
        assert Libro.objects.filter(propietario=user).count() == 3
        assert Libro.objects.filter(propietario=user, nombre='El Quijote').exists()
        assert Libro.objects.filter(propietario=user, nombre='1984').exists()
        assert Libro.objects.filter(propietario=user, nombre='Cien años de soledad').exists()
    
    def test_deteccion_duplicados_por_isbn(self, user):
        """Test detección de duplicados por ISBN"""
        # Crear libro existente con ISBN
        Libro.objects.create(
            nombre='Libro Original',
            autor='Autor Original',
            isbn='9781111111111',
            propietario=user
        )
        
        # Crear archivo Excel con libro duplicado por ISBN
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Autor', 'ISBN'])
        ws.append(['Libro Duplicado', 'Otro Autor', '9781111111111'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultados = procesar_excel_libros(excel_file, user)
        
        assert len(resultados['duplicados']) == 1
        assert resultados['duplicados'][0]['tipo'] == 'ISBN'
        assert len(resultados['libros_creados']) == 0
        # Verificar que no se creó el duplicado
        assert Libro.objects.filter(propietario=user).count() == 1
    
    def test_deteccion_duplicados_por_nombre_autor(self, user):
        """Test detección de duplicados por nombre+autor"""
        # Crear libro existente
        Libro.objects.create(
            nombre='Libro Original',
            autor='Autor Original',
            propietario=user
        )
        
        # Crear archivo Excel con libro duplicado por nombre+autor
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Autor'])
        ws.append(['Libro Original', 'Autor Original'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultados = procesar_excel_libros(excel_file, user)
        
        assert len(resultados['duplicados']) == 1
        assert resultados['duplicados'][0]['tipo'] == 'Nombre+Autor'
        assert len(resultados['libros_creados']) == 0
    
    def test_manejo_filas_con_errores(self, excel_file_con_errores, user):
        """Test manejo de filas con errores de validación"""
        resultados = procesar_excel_libros(excel_file_con_errores, user)
        
        assert resultados['total_procesado'] == 3
        assert len(resultados['errores']) == 2  # Dos filas con errores
        assert len(resultados['libros_creados']) == 1  # Una fila válida
        
        # Verificar que solo se creó el libro válido
        assert Libro.objects.filter(propietario=user).count() == 1
    
    def test_validacion_campos_obligatorios(self, user):
        """Test validación de campos obligatorios (nombre y autor)"""
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Autor'])
        ws.append(['', 'Autor Válido'])  # Nombre vacío
        ws.append(['Nombre Válido', ''])  # Autor vacío
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultados = procesar_excel_libros(excel_file, user)
        
        assert len(resultados['errores']) == 2
        assert any('Nombre' in error['mensaje'] for error in resultados['errores'])
        assert any('Autor' in error['mensaje'] for error in resultados['errores'])
    
    def test_creacion_libros_en_lote(self, excel_file_valido, user):
        """Test que los libros se crean en lote usando bulk_create"""
        resultados = procesar_excel_libros(excel_file_valido, user)
        
        # Verificar que todos los libros se crearon
        assert len(resultados['libros_creados']) == 3
        assert Libro.objects.filter(propietario=user).count() == 3
    
    def test_manejo_excepciones_archivo_corrupto(self, user):
        """Test manejo de excepciones con archivo corrupto"""
        # Crear archivo que no es un Excel válido
        archivo_invalido = BytesIO(b'No es un archivo Excel')
        archivo_invalido.name = 'archivo.xlsx'
        
        resultados = procesar_excel_libros(archivo_invalido, user)
        
        # Debe haber errores
        assert len(resultados['errores']) > 0
    
    def test_normalizacion_nombres_columnas(self, user):
        """Test que se normalizan correctamente los nombres de columnas"""
        wb = Workbook()
        ws = wb.active
        # Usar nombres de columnas con variaciones
        ws.append(['TÍTULO', 'AUTOR', 'EDITORIAL'])
        ws.append(['Libro Test', 'Autor Test', 'Editorial Test'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultados = procesar_excel_libros(excel_file, user)
        
        # Debe procesar correctamente a pesar de las variaciones
        assert len(resultados['libros_creados']) == 1
    
    def test_validacion_longitud_campos(self, user):
        """Test validación de longitud de campos"""
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Autor'])
        # Nombre muy largo (>255 caracteres)
        ws.append(['A' * 256, 'Autor Válido'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultados = procesar_excel_libros(excel_file, user)
        
        assert len(resultados['errores']) > 0
        assert any('excede' in error['mensaje'].lower() for error in resultados['errores'])
    
    def test_sin_columnas_obligatorias(self, user):
        """Test cuando no se encuentran las columnas obligatorias"""
        # Crear archivo Excel sin columnas obligatorias (usando nombres que no se reconocen)
        wb = Workbook()
        ws = wb.active
        ws.append(['Columna1', 'Columna2'])  # Nombres que no se reconocen como Nombre/Autor
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultados = procesar_excel_libros(excel_file, user)
        
        assert len(resultados['errores']) > 0
        assert any('obligatorias' in error['mensaje'].lower() for error in resultados['errores'])


class TestGenerarPlantillaExcel:
    """Tests para la función generar_plantilla_excel()"""
    
    def test_generacion_plantilla_valida(self):
        """Test que se genera una plantilla válida"""
        plantilla = generar_plantilla_excel()
        
        assert plantilla is not None
        assert isinstance(plantilla, BytesIO)
        
        # Verificar que es un archivo Excel válido
        from openpyxl import load_workbook
        plantilla.seek(0)
        wb = load_workbook(plantilla)
        ws = wb.active
        
        # Verificar encabezados
        primera_fila = [cell.value for cell in ws[1]]
        assert 'Nombre' in primera_fila
        assert 'Autor' in primera_fila
    
    def test_formato_correcto_archivo(self):
        """Test que el formato del archivo es correcto (.xlsx)"""
        plantilla = generar_plantilla_excel()
        
        assert plantilla is not None
        # Verificar que se puede leer como Excel
        from openpyxl import load_workbook
        plantilla.seek(0)
        wb = load_workbook(plantilla)
        assert wb is not None


class TestVistaCargarLibrosMasivo:
    """Tests para la vista cargar_libros_masivo()"""
    
    @pytest.mark.django_db
    def test_carga_exitosa_archivo_valido(self, client, user, excel_file_valido):
        """Test carga exitosa con archivo válido"""
        client.force_login(user)
        
        excel_file_valido.name = 'libros.xlsx'
        response = client.post('/libros/cargar-masivo/', {
            'archivo_excel': excel_file_valido
        })
        
        assert response.status_code == 200
        # Verificar que se crearon los libros
        assert Libro.objects.filter(propietario=user).count() == 3
    
    @pytest.mark.django_db
    def test_manejo_errores_validacion_formulario(self, client, user):
        """Test manejo de errores de validación del formulario"""
        client.force_login(user)
        
        # Enviar sin archivo
        response = client.post('/libros/cargar-masivo/', {})
        
        assert response.status_code == 200
        # Debe mostrar el formulario con errores
    
    @pytest.mark.django_db
    def test_mensajes_exito(self, client, user, excel_file_valido):
        """Test que se muestran mensajes de éxito"""
        from django.contrib.messages import get_messages
        client.force_login(user)
        
        excel_file_valido.name = 'libros.xlsx'
        response = client.post('/libros/cargar-masivo/', {
            'archivo_excel': excel_file_valido
        }, follow=True)
        
        messages = list(get_messages(response.wsgi_request))
        assert any('exitosamente' in str(m).lower() for m in messages)
    
    @pytest.mark.django_db
    def test_mensajes_warning_duplicados(self, client, user):
        """Test que se muestran mensajes de warning para duplicados"""
        from django.contrib.messages import get_messages
        client.force_login(user)
        
        # Crear libro original que será duplicado
        Libro.objects.create(
            nombre='Libro Original',
            autor='Autor Original',
            isbn='9781111111111',
            propietario=user
        )
        
        # Crear archivo Excel con libro duplicado
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Autor', 'ISBN'])
        ws.append(['Libro Duplicado ISBN', 'Otro Autor', '9781111111111'])  # Mismo ISBN
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'libros.xlsx'
        
        response = client.post('/libros/cargar-masivo/', {
            'archivo_excel': excel_file
        }, follow=True)
        
        messages = list(get_messages(response.wsgi_request))
        assert any('duplicado' in str(m).lower() for m in messages)
    
    @pytest.mark.django_db
    def test_mensajes_error(self, client, user, excel_file_con_errores):
        """Test que se muestran mensajes de error"""
        from django.contrib.messages import get_messages
        client.force_login(user)
        
        excel_file_con_errores.name = 'libros.xlsx'
        response = client.post('/libros/cargar-masivo/', {
            'archivo_excel': excel_file_con_errores
        }, follow=True)
        
        messages = list(get_messages(response.wsgi_request))
        assert any('error' in str(m).lower() for m in messages)
    
    @pytest.mark.django_db
    def test_renderizado_con_resultados(self, client, user, excel_file_valido):
        """Test que se renderiza correctamente con resultados"""
        client.force_login(user)
        
        excel_file_valido.name = 'libros.xlsx'
        response = client.post('/libros/cargar-masivo/', {
            'archivo_excel': excel_file_valido
        })
        
        assert response.status_code == 200
        assert 'resultados' in response.context
        assert response.context['mostrar_resultados'] is True


class TestVistaDescargarPlantillaExcel:
    """Tests para la vista descargar_plantilla_excel()"""
    
    @pytest.mark.django_db
    def test_descarga_plantilla(self, client, user):
        """Test que se puede descargar la plantilla"""
        client.force_login(user)
        
        response = client.get('/libros/descargar-plantilla/')
        
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'attachment' in response['Content-Disposition']
        assert 'plantilla_libros.xlsx' in response['Content-Disposition']
